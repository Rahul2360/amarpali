import openpyxl as xl
from openpyxl import Workbook
import json
workbook1=xl.load_workbook("Nyalazone_Task\\policy_search_excel.xlsx")
workbook2=xl.load_workbook("Nyalazone_Task\\T90_POL.xlsx")
worksheet1=workbook1.active
rows1=worksheet1.max_row
column1=worksheet1.max_column
sheet_name=workbook2.sheetnames
worbook3=Workbook()
search_no=[]
list_head=[]
for sheets in range(0,len(sheet_name)):
    sheet_select=workbook2.worksheets[sheets]
    rows2=sheet_select.max_row
    column2=sheet_select.max_column
    worksheet3=worbook3.create_sheet(sheet_name[sheets])
    # print(worksheet3)
    for i in range(2,rows1+1):
        policy_no=worksheet1.cell(row=i,column=1).value#9912 9913
        for j in range(2,rows2+1):
            list_data=[]
            sheet_policy_no=sheet_select.cell(row=j,column=1).value
            if sheet_policy_no==policy_no:
                for col in range(1,column2+1):
                    heading=sheet_select.cell(row=1,column=col).value
                    worksheet3.cell(row=1,column=col).value=heading
                    data_table=sheet_select.cell(row=j,column=col).value
                    list_data.append(data_table) 
                worksheet3.append(list_data)
worbook3.remove(worbook3.worksheets[0])
worbook3.save("Nyalazone_Task\\excel.xlsx")
            
   