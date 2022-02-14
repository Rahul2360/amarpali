from crypt import methods
from math import ceil
from select import select
from webbrowser import get
from flask import Flask,jsonify,request
import openpyxl as xl
from openpyxl import Workbook, load_workbook
cell_data_dict={}
final_data=[]
app=Flask(__name__)
app.config['JSON_SORT_KEYS'] = False
@app.route("/getdata")
def fetch_data():
    Workbook_load=xl.load_workbook("user_data.xlsx")
    sheet_select=Workbook_load.worksheets[0]
    max_rows=sheet_select.max_row
    max_cols=sheet_select.max_column
    for rows in range(2,max_rows+1):
        cell_data_dict={}
        for cols in range(1,max_cols+1):
            headings=sheet_select.cell(row=1,column=cols).value
            cell_data=sheet_select.cell(row=rows,column=cols).value
            cell_data_dict[headings]=cell_data
        final_data.append(cell_data_dict)
    return jsonify(final_data)
@app.route("/selectid=<int:n>")
def selectdata(n):

    Workbook_load=xl.load_workbook("user_data.xlsx")
    sheet_select=Workbook_load.worksheets[0]
    max_rows=sheet_select.max_row
    max_cols=sheet_select.max_column
    select_data_dict={}
    for rows in range(1,max_rows+1):
        value_id= sheet_select.cell(row=rows,column=1).value
        if value_id==n:
            for cols in range(1,max_cols):
                heading=sheet_select.cell(row=1,column=cols).value
                ceil_data=sheet_select.cell(row=rows,column=cols).value
                select_data_dict[heading]=ceil_data
    return jsonify(select_data_dict)
 
@app.route('/insertdata',methods=['POST'])
def insertdata():
        all_data_dict=request.get_json()
        data_list=[]
        Workbook_load=xl.load_workbook("user_data.xlsx")
        sheet_select=Workbook_load.worksheets[0]
        max_rows=sheet_select.max_row
        id=sheet_select.cell(row=max_rows,column=1).value
        data_list=[id+1]
        show_dict={"id":id+1}
        for key,value in all_data_dict.items():
            data_list.append(value)
            show_dict[key]=value
        data_list.append("Active")
        sheet_select.append(data_list)
        Workbook_load.save("user_data.xlsx")
        return jsonify(show_dict)

@app.route('/status_update',methods=['POST'])
def status():
    change_status=request.get_json() #json 
    value_id=change_status["id"]
    set_status=change_status["status"]
    Workbook_load=xl.load_workbook("user_data.xlsx")
    sheet_select=Workbook_load.worksheets[0]
    max_rows=sheet_select.max_row
    max_cols=sheet_select.max_column
    if  set_status=="Inactive" or set_status=="Active" or set_status=="inactive" or set_status=="active" :
        for row in range(1,max_rows+1):
             check_id=sheet_select.cell(row=row,column=1).value
             if check_id==value_id:
                 for col in range(1,max_cols+1):
                     check_col=str(sheet_select.cell(row=1,column=col).value) #Status
                     if check_col=="Status":
                         sheet_select.cell(row=row,column=col).value=set_status 
                         Workbook_load.save("user_data.xlsx")
                         return "status changed"  
        return "Id not found" 
    else:
       return "Invalid status"

@app.route("/update",methods=["POST"])
def update_value():
    my_data=request.get_json()
    id=my_data["id"]
    first_name=my_data["first_name"]
    last_name=my_data["last_name"]
    gender=my_data["gender"]
    password=my_data["password"]
    Workbook_load=xl.load_workbook("user_data.xlsx")
    sheet_select=Workbook_load.worksheets[0]
    max_row=sheet_select.max_row
    max_col=sheet_select.max_column
    for row in range(2,max_row+1):
        check_id=sheet_select.cell(row=row,column=1).value
        if check_id==id:
            password_old=sheet_select.cell(row=row,column=6).value
            if password!=password_old:
                sheet_select.cell(row=row,column=2).value=first_name
                sheet_select.cell(row=row,column=3).value=last_name
                sheet_select.cell(row=row,column=4).value=gender
                sheet_select.cell(row=row,column=6).value=password
                mail=sheet_select.cell(row=row,column=5).value
                Workbook_load.save("user_data.xlsx")
                return ({"id":id,"first_name":first_name,"last_name":last_name,"gender":gender,"email":mail,"password":password})
            return("password must be diffrent previously")
    return("id not found")    
if __name__=="__main__":
    app.run(debug=True)


