import xlrd
from openpyxl import Workbook
workbook=xlrd.open_workbook("/home/deepti/Downloads/T90_POL.xls")
total_sheets = len(workbook.sheet_names())
total_sheet=format(workbook.nsheets)
total_sheet=int(total_sheet)
policy_no=input("Enter Policy Number ")
sheetnames=workbook.sheet_names()
workbook_d = Workbook()

for sheet_no in range(total_sheet):
    select_sheet=workbook.sheet_by_index(sheet_no)
    worksheet_d= workbook_d.create_sheet(sheetnames[sheet_no])
    maximum_col=select_sheet.ncols
    maximum_row=select_sheet.nrows
  
    for j in range(maximum_row):
        all_data_list=[]
        if j==0:
            for p in range(maximum_col):
                all_data_list.append(select_sheet.cell_value(j,p))
        value=select_sheet.cell_value(j,0)

        if value==policy_no:   
            for k in range(maximum_col):
                all_data_list.append(select_sheet.cell_value(j,k))  
        worksheet_d.append(all_data_list)
          
workbook_d.save("/home/deepti/Downloads/combinedexcel.xlsx")     