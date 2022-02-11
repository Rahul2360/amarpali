import openpyxl as xl
import flask
from flask import Flask,jsonify
from openpyxl import Workbook

workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
worksheet_s=workbook_s.worksheets[0]
       
heading_list=[]
current_list=[]
final_dict={}

row_s = worksheet_s.max_row
col_s = worksheet_s.max_column
app = Flask(__name__)

@app.route("/")
def getthedata():
        temp=[]
        for r in range(1, col_s+1):
            check_value=worksheet_s.cell(row=1,column=r).value
            temp.append(check_value)
    
        for r in range(2, row_s+1):
            a=0
            new_dict = {}
            for c in range(1, col_s+1):
                check_value=worksheet_s.cell(row=r,column=c).value
                if check_value!=None:
                    new_dict[temp[a]]=check_value
                    a=a+1
            current_list.append(new_dict)
            final_dict["data"]=current_list
        print(final_dict)
    
        return final_dict
getthedata()

@app.route("/selectdata/<int:Id>")
def selectthedata(Id):
    print(type(Id))
    get_data= final_dict["data"][Id-1]
    print(get_data)
    return get_data
selectthedata(3)
if __name__=="__main__":
    app.run(debug=True)
    app.run(port=8000)