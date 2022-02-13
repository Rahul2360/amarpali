import openpyxl as xl
import flask
from flask import Flask,jsonify, request
from openpyxl import Workbook

current_list=[]

app = Flask(__name__)
@app.route("/")

def getthedata():
        workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
        worksheet_s=workbook_s.worksheets[0]
        row_s = worksheet_s.max_row
        col_s = worksheet_s.max_column
        heading_list=[]
        for r in range(1, col_s+1):
            check_value=worksheet_s.cell(row=1,column=r).value
            heading_list.append(check_value)
    
        for r in range(2, col_s+1):
            a=0
            new_dict = {}
            for c in range(1,col_s):
                check_value=worksheet_s.cell(row=r,column=c).value
                if check_value!=None:
                    new_dict[heading_list[a]]=check_value
                    a=a+1
            current_list.append(new_dict)
        return jsonify(current_list)
        #print(final_dict)
#getthedata()

@app.route("/selectdata/<int:n>")
def selectthedata(n):
    workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
    worksheet_s=workbook_s.worksheets[0]
    row_s = worksheet_s.max_row
    col_s = worksheet_s.max_column
    for i in range(2,row_s+1):
       fetch_data=[]
       my_dict={}
       id=worksheet_s.cell(row=i,column=1).value
       if id==n:
           for j in range(1,col_s+1):
               key=worksheet_s.cell(row=1,column=j).value
               data=worksheet_s.cell(row=i,column=j).value
               my_dict[key]=data
           return my_dict
    return "Invalid Id/not found"

@app.route("/insert",methods=['POST'])
def insert_data():
    my_data=request.get_json()
    first_name=my_data["first_name"]
    last_name=my_data["last_name"]
    gender=my_data["gender"]
    email=my_data["email"]
    password=my_data["password"]
    workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
    worksheet_s=workbook_s.worksheets[0]
    row_s = worksheet_s.max_row
    c=worksheet_s.cell(row=row_s,column=1).value
    newlist=[c+1,first_name,last_name,gender,email,password,"Active"]
    worksheet_s.append(newlist)
    workbook_s.save("/home/deeptia/Downloads/user_data.xlsx")
    return jsonify({"Id":newlist[0],"first_name":newlist[1],"last_name":newlist[2],"gender":newlist[3],"email":newlist[4],"password":newlist[5]})

@app.route("/status",methods=['POST'])
def update_status():
     change_status=request.get_json()
     id=change_status["id"]
     status=change_status["status"]
     workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
     worksheet_s=workbook_s.worksheets[0]
     row_s = worksheet_s.max_row
     col_s = worksheet_s.max_column
     if status=="inactive" or status=="Inactive" or status=="Active" or status=="active" :
         for i in range(1,row_s+1):
             check=worksheet_s.cell(row=i,column=1).value
             if check==id:
                for j in range(1,col_s+1):
                     check_col=worksheet_s.cell(row=1,column=j).value
                     if check_col=="Status":
                        worksheet_s.cell(row=i,column=j).value=status
                        workbook_s.save("/home/deepti/Downloads/user_data.xlsx")
                        return "Status is changed"
             
                return "Id not available or found"
     else:
        return "Invalid status"

@app.route("/update",methods=['POST'])
def update_data():
   my_data=request.get_json()
   id=my_data["id"]
   first_name=my_data["first_name"]
   last_name=my_data["last_name"]
   gender=my_data["gender"]
   password=my_data["password"]
   workbook_s= xl.load_workbook("/home/deepti/Downloads/user_data.xlsx")
   worksheet_s=workbook_s.worksheets[0]
   row_s = worksheet_s.max_row
   for i in range(1,row_s):
       check=worksheet_s.cell(row=i,column=1).value
       if check==id:
           password_old=worksheet_s.cell(row=i,column=6).value
           if password != password_old:
               worksheet_s.cell(row=i,column=2).value=first_name
               worksheet_s.cell(row=i,column=3).value=last_name
               worksheet_s.cell(row=i,column=4).value=gender
               worksheet_s.cell(row=i,column=6).value=password
               email=worksheet_s.cell(row=i,column=5).value
               workbook_s.save("/home/deepti/Downloads/user_data.xlsx")
               return ({"id":id,"first_name":first_name,"last_name":last_name,"gender":gender,"email":email,"password":password})
   return "Password is used previously.Please use a different password."

if __name__=="__main__":
    app.run(debug=True)
    app.run(port=8000)