import openpyxl as xl
from flask import Flask,jsonify,request
cell_dict={}
final_data=[]
app=Flask(__name__)
app.config['JSON_SORT_KEYS'] = False
@app.route("/")
def fetch_excel():
    wb_s=xl.load_workbook("user_data.xlsx")
    ws_s=wb_s.worksheets[0]
    max_rows=ws_s.max_row
    max_cols=ws_s.max_column
    # print(max_rows)
    for rows in range(2,max_rows+1):
        cell_dict={}
        for cols in range(1,max_cols+1):
                head_data=ws_s.cell(row=1,column=cols).value
                cell_data=ws_s.cell(row=rows,column=cols).value
                cell_dict[head_data]=cell_data
        final_data.append(cell_dict)      
    return jsonify(final_data) 

@app.route('/id=<int:n>')
def datawithid(n):
    #return jsonify(final_data[n-1]) #here we can only use this line instead of whole bottomn code if id are in serial 1,2,3,4 format
    # for i in range(0,len(final_data)):
    #     if final_data[i]["Id"]==n:
    #         return jsonify(final_data[i])

    workbook_s=xl.load_workbook("user_data.xlsx")
    worksheet_s=workbook_s.worksheets[0]
    cols=worksheet_s.max_column
    rows=worksheet_s.max_row
    user_data_dict={}
    for j in range(1,rows+1):
        id_values=worksheet_s.cell(row=j,column=1).value
        if id_values==n:
            for k in range(1,cols+1):
                headings=worksheet_s.cell(row=1,column=k).value
                user_data=worksheet_s.cell(row=j,column=k).value
                user_data_dict[headings]=user_data
            return jsonify(user_data_dict) 
    return ("<script>window.alert(`Invalid Id`)</script>")

@app.route('/insert',methods=['POST'])
def insertdata():
    data_dict=request.get_json()
    if len(data_dict)!=5:
        return("invalid number of attributes")
    elif len(data_dict)==5:
        data_list=[]
        wb_s=xl.load_workbook("user_data.xlsx")
        ws_s=wb_s.worksheets[0]
        row_s=ws_s.max_row
        id=ws_s.cell(row=row_s,column=1).value
        data_list=[id+1]
        show_dict={"id":id+1}
        for key,value in data_dict.items():
            data_list.append(value)
            show_dict[key]=value
        data_list.append("Active")
        ws_s.append(data_list)
        wb_s.save("user_data.xlsx")
        return jsonify(show_dict)
#    return jsonify({"id":data_list[0],"first_name":data_list[1],"last_name":data_list[2],"gender":data_list[3],"email":data_list[4],"password":data_list[5]})

@app.route('/status_update',methods=['POST'])
def update_status():
    change_status=request.get_json()
    if len(change_status)!=2:
        return("invalid number of attributes")
    elif len(change_status)==2:
        id=change_status["id"]
        status=change_status["status"]
        wb_s=xl.load_workbook("user_data.xlsx")
        ws_s=wb_s.worksheets[0]
        row_s=ws_s.max_row
        col_s=ws_s.max_column
        if  status=="Inactive" or status=="Active" or status=="inactive" or status=="active" :
            for i in range(2,row_s+1):
                check=ws_s.cell(row=i,column=1).value
                if id==check:
                    for j in range(1,col_s+1):
                        find_status=str(ws_s.cell(row=1,column=j).value)
                    #print(find_status)
                        if find_status=="Status":
                            ws_s.cell(row=i,column=j).value=status 
                            wb_s.save("user_data.xlsx")   
                            return ("id="+str(id)+" status_changed to "+str(status))  
          
            return ('<h1 style="color:red">Id Not Found</h1>') 
        else:
            return ('Invalid Status')

@app.route("/other_updates",methods=["POST"])
def update_others():
    data_dict=request.get_json()
    if len(data_dict)!=5:
        return("invalid number of attributes")
    elif len(data_dict)==5:
        response_dict={}
        data_list=[]
        for key in data_dict:
            data_list.append(data_dict[key])
        wb_s=xl.load_workbook("user_data.xlsx")
        ws_s=wb_s.worksheets[0]
        row_s=ws_s.max_row
        col_s=ws_s.max_column
        for i in range(2,row_s+1):
            match=ws_s.cell(row=i,column=1).value
            old_password=ws_s.cell(row=i,column=6).value 
            if match==data_list[0]:
                old_password=ws_s.cell(row=i,column=6).value 
                if data_list[4]!=old_password:  
                    ws_s.cell(row=i,column=2).value=data_list[1]
                    ws_s.cell(row=i,column=3).value=data_list[2]
                    ws_s.cell(row=i,column=4).value=data_list[3]
                    ws_s.cell(row=i,column=6).value=data_list[4]
                    for j in range(1,col_s):
                        head=ws_s.cell(row=1,column=j).value
                        response_dict[head]=ws_s.cell(row=i,column=j).value
                    wb_s.save("user_data.xlsx")  
                    return(response_dict)
        
                return("Password used previously.Please insert a different password.")    
        return('<h1 style="color:red">Id Not Found</h1>')      
    
if __name__=="__main__":
    app.run(debug=True)