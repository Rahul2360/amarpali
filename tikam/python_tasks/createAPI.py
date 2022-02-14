from flask import Flask,jsonify,request
import openpyxl as xl
from openpyxl import Workbook
 
current_list=[]
app = Flask(__name__)
@app.route("/")
def getdata():
      
    workbook_s= xl.load_workbook("user_data.xlsx")
    worksheet_s=workbook_s.worksheets[0]
    row = worksheet_s.max_row
    col = worksheet_s.max_column
    head_list=[]
    for k in range(1, col+1):
        cellvalue=worksheet_s.cell(row=1,column=k).value
        head_list.append(cellvalue)
        
    for r in range(2, row+1):
            a=0
            new_dict = {}
            for c in range(1,col):
                cellvalue=worksheet_s.cell(row=r,column=c).value
                if cellvalue!=None:
                   new_dict[head_list[a]]=cellvalue
                   a=a+1
            current_list.append(new_dict)
    return jsonify(current_list)
 
@app.route("/id=<int:n>")
def selectdata(n):
   workbook_s= xl.load_workbook("user_data.xlsx")
   worksheet_s=workbook_s.worksheets[0]
   row_s = worksheet_s.max_row
   col_s = worksheet_s.max_column
   for i in range(2,row_s+1):
       fetch_data=[]
       my_dict={}
       id=worksheet_s.cell(row=i,column=1).value
       if id==n:
           for j in range(1,col_s+1):
               key=worksheet_s.cell(row=1,column=j).value
               data=worksheet_s.cell(row=i,column=j).value
               my_dict[key]=data
           return my_dict
   return "invalid Id"

@app.route("/insert",methods=['POST'])
def insert_data():
   data=request.get_json()
   first_name=data["first_name"]
   last_name=data["last_name"]
   gender=data["gender"]
   email=data["email"]
   password=data["password"]
   workbook_s= xl.load_workbook("user_data.xlsx")
   worksheet_s=workbook_s.worksheets[0]
   row_s = worksheet_s.max_row
 
   id=worksheet_s.cell(row=row_s,column=1).value
   insert_list=[id+1,first_name,last_name,gender,email,password,"Active"]
   worksheet_s.append(insert_list)
 
   workbook_s.save("user_data.xlsx")
   return jsonify({"Id":insert_list[0],"first_name":insert_list[1],"last_name":insert_list[2],"gender":insert_list[3],"email":insert_list[4],"password":insert_list[5]})

@app.route("/status",methods=['POST'])
def change_status():
    status_data=request.get_json()
    id=status_data["id"]
    status=status_data["status"]
    workbook_s=xl.load_workbook("user_data.xlsx")
    worksheet_s=workbook_s.worksheets[0]
    max_row=worksheet_s.max_row
    max_col=worksheet_s.max_column
    if status=="Active" or status=="Inactive" or status=="active" or status=="inactive":
        for i in range(1,max_row):
            id_value=worksheet_s.cell(row=i,column=1).value
            if id==id_value:
                for j in range(1,max_col):
                    row_value=worksheet_s.cell(row=1,column=j).value
                    print(row_value)
                    if row_value=="status":
                        worksheet_s.cell(row=i,column=j).value=status
                        workbook_s.save("user_data.xlsx")
                        return "status changed"
                
    else:
        return "Invalid status"

@app.route("/update",methods=['POST'])
def update_data():
   data=request.get_json()
   id=data["id"]
   first_name=data["first_name"]
   last_name=data["last_name"]
   gender=data["gender"]
   password=data["password"]
   workbook_s= xl.load_workbook("user_data.xlsx")
   worksheet_s=workbook_s.worksheets[0]
   row_s = worksheet_s.max_row
   for i in range(1,row_s):
       check=worksheet_s.cell(row=i,column=1).value
       if check==id:
           password_old=worksheet_s.cell(row=i,column=6).value
           if password != password_old:
               worksheet_s.cell(row=i,column=2).value=first_name
               worksheet_s.cell(row=i,column=3).value=last_name
               worksheet_s.cell(row=i,column=4).value=gender
               worksheet_s.cell(row=i,column=6).value=password
               email=worksheet_s.cell(row=i,column=5).value
               workbook_s.save("user_data.xlsx")
               return ({"id":id,"first_name":first_name,"last_name":last_name,"gender":gender,"email":email,"password":password})
   return "Password is used previously.Please use a different password."
if __name__=="__main__":
   app.run(debug=True)
   app.run(port=8000)


  